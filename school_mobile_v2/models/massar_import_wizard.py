import base64
import io
from odoo import models, fields, api
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class SchoolMassarImportWizard(models.TransientModel):
    _name = 'school.massar.import.wizard'
    _description = 'Assistant d\'Importation des Notes MASSAR (Excel)'

    file = fields.Binary(string='Fichier Excel MASSAR (.xlsx)', required=True)
    filename = fields.Char(string='Nom du fichier')
    level_id = fields.Many2one('school.level', string='Niveau / Classe (Optionnel - Détecté auto)')
    subject_id = fields.Many2one('school.subject', string='Matière (Optionnel - Détectée auto)')
    semester = fields.Selection([
        ('S1', 'Semestre 1 (الدورة الأولى)'),
        ('S2', 'Semestre 2 (الدورة الثانية)')
    ], string='Semestre (Optionnel - Détecté auto)')
    cc_type = fields.Selection([
        ('cc1', 'CC1 (الفرض الأول)'),
        ('cc2', 'CC2 (الفرض الثاني)')
    ], string='Type de Contrôle (Optionnel - Détecté auto)')

    @api.model
    def _register_hook(self):
        super()._register_hook()
        try:
            model = self.env['ir.model'].sudo().search([('model', '=', 'school.massar.import.wizard')], limit=1)
            if model:
                access = self.env['ir.model.access'].sudo().search([('model_id', '=', model.id)], limit=1)
                if not access:
                    self.env['ir.model.access'].sudo().create({
                        'name': 'school.massar.import.wizard access',
                        'model_id': model.id,
                        'perm_read': True,
                        'perm_write': True,
                        'perm_create': True,
                        'perm_unlink': True,
                    })
        except Exception:
            pass

    def action_import_notes(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError("La bibliothèque python 'openpyxl' est requise sur le serveur pour importer les fichiers Excel.")

        if not self.file:
            raise UserError("Veuillez sélectionner un fichier Excel MASSAR (.xlsx).")

        file_data = base64.b64decode(self.file)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
        except Exception as e:
            raise UserError(f"Impossible de lire le fichier Excel : {e}")

        sheet = wb.active

        detected_subject = ''
        detected_level = ''
        detected_teacher = ''
        detected_semester = self.semester or 'S1'
        detected_cc = self.cc_type or 'cc1'

        # 1. Scanner les 16 premières lignes d'entête
        for r in range(1, min(17, sheet.max_row + 1)):
            for c in range(1, min(20, sheet.max_column + 1)):
                val = str(sheet.cell(row=r, column=c).value or '').strip()
                if val == 'المادة' or val == 'Matière':
                    for k in range(c + 1, min(c + 5, sheet.max_column + 1)):
                        v = str(sheet.cell(row=r, column=k).value or '').strip()
                        if v:
                            detected_subject = v
                            break
                elif val in ['الاستاذ', 'الأستاذ', 'Professeur', 'Enseignant']:
                    for k in range(c + 1, min(c + 5, sheet.max_column + 1)):
                        v = str(sheet.cell(row=r, column=k).value or '').strip()
                        if v:
                            detected_teacher = v
                            break
                elif 'القسم' in val or 'المستوى' in val or 'Classe' in val:
                    for k in range(c + 1, min(c + 5, sheet.max_column + 1)):
                        v = str(sheet.cell(row=r, column=k).value or '').strip()
                        if v:
                            if not detected_level or len(v) < len(detected_level):
                                detected_level = v
                            break
                elif 'الدورة' in val or 'Semestre' in val:
                    for k in range(c + 1, min(c + 5, sheet.max_column + 1)):
                        v = str(sheet.cell(row=r, column=k).value or '').strip()
                        if 'ثانية' in v or '2' in v or 'S2' in v:
                            detected_semester = 'S2'
                        elif 'أولى' in v or '1' in v or 'S1' in v:
                            detected_semester = 'S1'
                elif 'نقط' in val or 'الفرض' in val or 'Contrôle' in val:
                    for k in range(c + 1, min(c + 5, sheet.max_column + 1)):
                        v = str(sheet.cell(row=r, column=k).value or '').strip()
                        if 'ثاني' in v or '2' in v or 'CC2' in v:
                            detected_cc = 'cc2'
                        elif 'أول' in v or '1' in v or 'CC1' in v:
                            detected_cc = 'cc1'

        # 2. Trouver la matière dans Odoo
        target_subject = self.subject_id
        if not target_subject and detected_subject:
            target_subject = self.env['school.subject'].search([
                '|', ('name', 'ilike', detected_subject),
                ('code', 'ilike', detected_subject)
            ], limit=1)
        if not target_subject:
            # Recherche élargie
            if 'عرب' in detected_subject:
                target_subject = self.env['school.subject'].search([('name', 'ilike', 'عربية')], limit=1)
            elif 'فرنس' in detected_subject or 'fr' in detected_subject.lower():
                target_subject = self.env['school.subject'].search([('name', 'ilike', 'français')], limit=1)
            elif 'رياض' in detected_subject or 'math' in detected_subject.lower():
                target_subject = self.env['school.subject'].search([('name', 'ilike', 'math')], limit=1)
            elif 'إسلام' in detected_subject or 'islam' in detected_subject.lower():
                target_subject = self.env['school.subject'].search([('name', 'ilike', 'إسلام')], limit=1)
            elif 'اجتماع' in detected_subject or 'hist' in detected_subject.lower():
                target_subject = self.env['school.subject'].search([('name', 'ilike', 'اجتماع')], limit=1)
            elif 'علم' in detected_subject or 'sci' in detected_subject.lower():
                target_subject = self.env['school.subject'].search([('name', 'ilike', 'علم')], limit=1)
            elif 'فن' in detected_subject or 'art' in detected_subject.lower():
                target_subject = self.env['school.subject'].search([('name', 'ilike', 'فن')], limit=1)

        # 3. Créer ou associer le professeur si détecté
        teacher_rec = False
        if detected_teacher:
            teacher_rec = self.env['school.teacher'].search([('name', '=', detected_teacher)], limit=1)
            if not teacher_rec:
                teacher_rec = self.env['school.teacher'].create({
                    'name': detected_teacher,
                    'subject': target_subject.name if target_subject else False,
                    'subject_ids': [(4, target_subject.id)] if target_subject else False
                })
            else:
                if target_subject and target_subject not in teacher_rec.subject_ids:
                    teacher_rec.write({'subject_ids': [(4, target_subject.id)]})

        # 4. Trouver la ligne d'entête des colonnes élèves (رقم التلميذ)
        header_row = -1
        massar_col = -1
        name_col = -1

        for r in range(12, min(22, sheet.max_row + 1)):
            for c in range(1, min(10, sheet.max_column + 1)):
                val = str(sheet.cell(row=r, column=c).value or '').strip()
                if 'رقم' in val and 'تلميذ' in val:
                    header_row = r
                    massar_col = c
                elif 'إسم' in val and 'تلميذ' in val:
                    name_col = c

        if header_row == -1 or massar_col == -1:
            raise UserError("Impossible de trouver la colonne 'رقم التلميذ' dans le fichier Excel.")

        # 5. Identifier les colonnes de sous-matières / composantes
        component_cols = []
        for c in range(massar_col + 2, sheet.max_column + 1):
            comp_name = str(sheet.cell(row=header_row, column=c).value or '').strip().replace('\r', '').replace('\n', ' ')
            if comp_name and not any(x in comp_name for x in ['ملاحظات', 'تاريخ', 'ID', 'الأستاذ', '-']):
                # Trouver la sous-matière correspondante
                sub_subj = False
                if target_subject:
                    sub_subj = self.env['school.sub.subject'].search([
                        ('subject_id', '=', target_subject.id),
                        ('name', 'ilike', comp_name.split('(')[0].strip())
                    ], limit=1)
                component_cols.append({
                    'col': c,
                    'name': comp_name,
                    'sub_subject_id': sub_subj.id if sub_subj else False
                })

        # 6. Parcourir les lignes d'élèves et enregistrer les notes
        imported_count = 0
        updated_students = set()

        for r in range(header_row + 2, sheet.max_row + 1):
            massar_val = str(sheet.cell(row=r, column=massar_col).value or '').strip()
            student_name = str(sheet.cell(row=r, column=name_col).value or '').strip() if name_col != -1 else ''

            if not massar_val and not student_name:
                continue

            # Trouver l'élève par Code Massar ou Nom
            student = False
            if massar_val:
                student = self.env['school.student'].search([('massar_number', '=', massar_val)], limit=1)
            if not student and student_name:
                student = self.env['school.student'].search([('name', '=', student_name)], limit=1)

            if not student:
                continue

            # Extraire les notes de la ligne
            marks = []
            for comp in component_cols:
                mark_val = sheet.cell(row=r, column=comp['col']).value
                if mark_val is not None and mark_val != '':
                    try:
                        fmark = float(mark_val)
                        marks.append(fmark)
                    except (ValueError, TypeError):
                        pass

            # Calculer la moyenne des composantes si présentes
            calculated_mark = 0.0
            if marks:
                calculated_mark = sum(marks) / len(marks)

            if calculated_mark > 0 or marks:
                # Trouver ou créer l'enregistrement school.grade pour cet élève et cette matière
                grade_domain = [
                    ('student_id', '=', student.id),
                    ('semester', '=', detected_semester)
                ]
                if target_subject:
                    grade_domain.append(('subject_id', '=', target_subject.id))

                grade_rec = self.env['school.grade'].search(grade_domain, limit=1)

                vals = {
                    'student_id': student.id,
                    'subject_id': target_subject.id if target_subject else False,
                    'subject': target_subject.name if target_subject else detected_subject,
                    'semester': detected_semester,
                }

                if detected_cc == 'cc1':
                    vals['cc1'] = calculated_mark
                elif detected_cc == 'cc2':
                    vals['cc2'] = calculated_mark

                # Si les deux CC existent, mettre à jour la note finale
                if grade_rec:
                    if detected_cc == 'cc1':
                        vals['final_mark'] = (calculated_mark + grade_rec.cc2) / 2 if grade_rec.cc2 > 0 else calculated_mark
                    else:
                        vals['final_mark'] = (grade_rec.cc1 + calculated_mark) / 2 if grade_rec.cc1 > 0 else calculated_mark
                    grade_rec.write(vals)
                else:
                    vals['final_mark'] = calculated_mark
                    self.env['school.grade'].create(vals)

                imported_count += 1
                updated_students.add(student.id)

        # 7. Notification de succès
        subject_label = target_subject.name if target_subject else detected_subject
        teacher_label = f" (Prof: {detected_teacher})" if detected_teacher else ""
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': '🎉 Importation MASSAR Réussie !',
                'message': f"{imported_count} notes importées avec succès pour {len(updated_students)} élèves.\nMatière : {subject_label}{teacher_label} - Semestre : {detected_semester} ({detected_cc.upper()})",
                'sticky': False,
                'type': 'success',
            }
        }
